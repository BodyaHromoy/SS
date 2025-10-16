import json
import random
import string
from datetime import timedelta
import datetime
import pytz
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models.functions import Cast
import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView, LogoutView
from django.core.mail import send_mail
from django.db.models import Count, Avg
from django.http import HttpResponse
from django.shortcuts import redirect
from django.db.models import F, FloatField
import paho.mqtt.client as mqtt
from django.db import transaction, DataError, IntegrityError
import logging
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from django.http import JsonResponse
from django.contrib.auth.decorators import user_passes_test
from .decorators.auth_decorators import staff_required
from .forms.auth_form import CustomAuthenticationForm
from .forms.forms import ReportFilterForm, CourierCreationForm, LogicCreationForm, CabinetSettingsForm, \
    SettingsForSettingsForm
from .models import *
from django.utils import timezone
import re
import requests
from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
from django.conf import settings
from django.shortcuts import get_object_or_404, render
from .models import Cabinet
from django.utils.dateparse import parse_date

logger = logging.getLogger(__name__)


def is_courier(user):
    return user.is_authenticated and user.role == 'courier' or 'engineer'


def is_logistician(user):
    return user.is_authenticated and user.role == 'logistician' or 'engineer'


def is_engineer(user):
    return user.is_authenticated and user.role == 'engineer'


def is_regional_manager(user):
    return user.is_authenticated and user.role == 'regional_manager' or 'engineer'


def map_view(request):
    cabinets = Cabinet.objects.all()
    data = []
    for cabinet in cabinets:
        cells = Cell.objects.filter(cabinet_id=cabinet)
        status_counts = cells.values("status").annotate(count=models.Count("status"))
        status_dict = {status["status"]: status["count"] for status in status_counts}

        data.append({
            "shkaf_id": cabinet.shkaf_id,
            "latitude": cabinet.latitude,
            "longitude": cabinet.longitude,
            "status_counts": status_dict,
        })

    return render(request, 'ss_main/map_page.html', {"cabinets": data})


def get_cabinets(request):
    cabinets = Cabinet.objects.all()
    data = []
    for cabinet in cabinets:
        # Получаем все ячейки для этого шкафа
        cells = Cell.objects.filter(cabinet_id=cabinet)

        # Подсчитываем количество каждого статуса
        status_counts = cells.values("status").annotate(count=models.Count("status"))

        # Преобразуем список статусов в словарь для удобства
        status_dict = {status["status"]: status["count"] for status in status_counts}

        # Добавляем информацию о шкафе и статусах
        data.append({
            "shkaf_id": cabinet.shkaf_id,
            "latitude": cabinet.latitude,
            "longitude": cabinet.longitude,
            "status_counts": status_dict,  # Добавляем словарь с количеством статусов
            "city": cabinet.city.city_name,
            "zone": cabinet.zone.zone_name,
            "location": cabinet.location,
            "extrainf": cabinet.extra_inf,
            "street": cabinet.street,
        })

    return JsonResponse({"cabinets": data})



@user_passes_test(is_engineer)
def send_command(request):
    if request.method == 'POST':
        cabinet_id = request.POST.get('cabinet_id')
        endpoint_id = request.POST.get('endpoint_id')
        cmd_number = request.POST.get('cmd_number')

        try:
            if cmd_number == "del":
                deleted_count, _ = Cell.objects.filter(cabinet_id__shkaf_id=cabinet_id, endpointid=endpoint_id).delete()
                if deleted_count > 0:
                    return JsonResponse({"success": True, "message": "Запись успешно удалена!"})
                else:
                    return JsonResponse({"success": False, "message": "Запись не найдена для удаления."})

            record = Cell.objects.get(cabinet_id__shkaf_id=cabinet_id, endpointid=endpoint_id)

            if cmd_number == '1':
                record.is_error = True
                Marked.objects.create(sn=record.sn)
            elif cmd_number == '0':
                record.is_error = False
                Marked.objects.filter(sn=record.sn).delete()
            record.save()

            json_data = {
                "Type": "cmd",
                "StationID": int(record.cabinet_id_id),
                "EndpointID": int(record.endpointid),
                "CMD": int(cmd_number),
                "SN": record.sn
            }

            client = mqtt.Client()
            client.connect("192.168.1.98", 1883, 60)
            client.publish("test/back", json.dumps(json_data))
            client.disconnect()

            return JsonResponse({"success": True, "message": "Команда отправлена успешно!"})
        except Cell.DoesNotExist:
            return JsonResponse({"success": False, "message": "Ячейка не найдена."})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Ошибка при обработке команды: {str(e)}"})
    return JsonResponse({"success": False, "message": "Неверный запрос."})


@user_passes_test(is_engineer)
def new_eng(request):
    cabinets = list(Cabinet.objects.all())

    for cabinet in cabinets:
        if cabinet.iot_imei_locker:
            rssi, door_state, _ , _= parse_device_status(cabinet.iot_imei_locker)
            cabinet.rssi = rssi
            cabinet.door_state = door_state
        else:
            cabinet.rssi = '-'
            cabinet.door_state = None

    return render(request, "ss_main/new_eng.html", {
        "cabinets": cabinets,
        "cities": City.objects.all(),
        "zones": Zone.objects.all(),
        "vendors": Vendor.objects.all(),
    })



@user_passes_test(is_engineer)
def new_eng_cabinet_detail(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)
    cells = Cell.objects.filter(cabinet_id=cabinet).order_by('endpointid')
    return render(request, 'ss_main/new_eng_cabinet_detail.html', {
        'cabinet': cabinet,
        'cells': cells
    })


@user_passes_test(is_engineer)
def cabinet_card(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)
    data = {
        "city": cabinet.city.city_name if cabinet.city else "",
        "zone": cabinet.zone.zone_name if cabinet.zone else "",
        "street": cabinet.street,
        "extra_inf": cabinet.extra_inf,
        "vendor": cabinet.device_vendor,
        "qr": cabinet.qr,
        "n_inventar": cabinet.n_inventar,
        "shkaf_id": cabinet.shkaf_id,
        "capacity": cabinet.capacity,
        "buffer": cabinet.buffer,
        "energy_counter_sn": cabinet.energy_counter_sn,
        "iot_imei_rpi": cabinet.device_id,
        "mobile_n_rpi": cabinet.mobile_n_rpi,
        "iot_imei_locker": cabinet.iot_imei_locker,
        "mobile_n_locker": cabinet.mobile_n_locker,
        "readable_name": cabinet.readable_name,
    }
    data["all_cities"] = list(City.objects.values_list("city_name", flat=True).order_by("city_name"))
    data["all_zones"] = list(Zone.objects.values_list("zone_name", flat=True).order_by("zone_name"))
    return JsonResponse(data)


logger = logging.getLogger(__name__)


@require_POST
@user_passes_test(is_engineer)
def save_cabinet_card(request):
    try:
        shkaf_id = request.POST.get('shkaf_id')
        if not shkaf_id:
            return JsonResponse({'success': False, 'error': 'shkaf_id not provided'}, status=400)

        cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)

        # --- вспомогательные функции ---
        def to_str(val):
            if val is None:
                return None
            if hasattr(val, 'city_name'):
                return val.city_name
            if hasattr(val, 'zone_name'):
                return val.zone_name
            return str(val)

        def normalize(val):
            if val is None:
                return ""
            return str(val).strip()

        # старые значения
        old_values = {f: to_str(getattr(cabinet, f, None)) for f in [
            'city', 'zone', 'street', 'extra_inf', 'device_vendor',
            'qr', 'n_inventar', 'capacity', 'buffer', 'energy_counter_sn',
            'device_id', 'mobile_n_rpi', 'iot_imei_locker', 'mobile_n_locker',
            'readable_name', 'door_state'
        ]}

        # новые значения
        city_name = request.POST.get('city')
        cabinet.city = City.objects.filter(city_name=city_name).first() if city_name else None
        zone_name = request.POST.get('zone')
        cabinet.zone = Zone.objects.filter(zone_name=zone_name).first() if zone_name else None
        cabinet.street = request.POST.get('street') or ''
        cabinet.extra_inf = request.POST.get('extra_inf') or ''
        cabinet.device_vendor = request.POST.get('vendor') or ''
        cabinet.qr = request.POST.get('qr') or ''
        cabinet.n_inventar = request.POST.get('n_inventar') or ''
        cabinet.capacity = request.POST.get('capacity') or None
        cabinet.buffer = request.POST.get('buffer') or ''
        cabinet.energy_counter_sn = request.POST.get('energy_counter_sn') or ''
        cabinet.device_id = request.POST.get('iot_imei_rpi') or ''
        cabinet.mobile_n_rpi = request.POST.get('mobile_n_rpi') or ''
        cabinet.iot_imei_locker = request.POST.get('iot_imei_locker') or ''
        cabinet.mobile_n_locker = request.POST.get('mobile_n_locker') or ''
        cabinet.readable_name = request.POST.get('readable_name') or ''

        with transaction.atomic():
            cabinet.save()

            new_values = {f: to_str(getattr(cabinet, f, None)) for f in old_values.keys()}

            changes = [
                f"{key}: {old_values[key]} → {new_values[key]}"
                for key in old_values if normalize(old_values[key]) != normalize(new_values[key])
            ]

            if changes:
                action = "; ".join(changes)
                if len(action) > 250:
                    action = action[:250] + "…"

                Eventlogger.objects.create(
                    user=request.user.username if request.user.is_authenticated else "anonymous",
                    shkaf_id=cabinet.shkaf_id or '',
                    login=request.user.username if request.user.is_authenticated else '',
                    action=action,
                    time=timezone.now(),
                    door_state_before=old_values.get('door_state') in ('True', 'true', True)
                )

        return JsonResponse({'success': True})

    except Exception as exc:
        logger.exception("save_cabinet_card failed")
        return JsonResponse({'success': False, 'error': str(exc)})





@require_POST
@user_passes_test(is_engineer)
def save_cabinet(request):
    data = request.POST
    try:
        shkaf_id = data.get('shkaf_id')
        cabinet, created = Cabinet.objects.update_or_create(
            shkaf_id=shkaf_id,
            defaults={
                'city_id': data.get('city'),
                'zone_id': data.get('zone'),
                'vendor_id': data.get('vendor'),
                'street': data.get('street'),
                'location': data.get('location'),
                'extra_inf': data.get('extra_inf', ''),
                'latitude': data.get('latitude') or None,
                'longitude': data.get('longitude') or None,
                'device_id': data.get('device_id', ''),
                'capacity': data.get('capacity', ''),
                'buffer': data.get('buffer', ''),
                'sense': data.get('sense', ''),
                'rssi': data.get('rssi', ''),
                'sticker': data.get('sticker', ''),
                'fire_allert': data.get('fire_allert') == 'true',
                'door_state': data.get('door_state') == 'true',
            }
        )

        settings, _ = Cabinet_settings_for_auto_marking.objects.update_or_create(
            cabinet_id=cabinet,
            defaults={
                'settings_for': shkaf_id,
                'sn_error': data.get('sn_error') == 'true',
                'year_of_manufacture': data.get('year_of_manufacture'),
                'max_cycle_times': data.get('max_cycle_times'),
                'vid': data.get('vid'),
                'sw_ver': data.get('sw_ver'),
                'critical_temp': data.get('critical_temp') or None,
                'lock_status': data.get('lock_status') == 'true',
                'smoke_status': data.get('smoke_status') == 'true',
                'temp_inside': data.get('temp_inside', ''),
                'fan_status': data.get('fan_status') == 'true',
                'mains_voltage': data.get('mains_voltage'),
                'reserve_voltage': data.get('reserve_voltage')
            }
        )

        Settings_for_settings.objects.update_or_create(
            settings_for=settings,
            defaults={
                'sn_error': data.get('sfs_sn_error') == 'true',
                'year_of_manufacture': data.get('sfs_year_of_manufacture') == 'true',
                'max_cycle_times': data.get('sfs_max_cycle_times') == 'true',
                'vid': data.get('sfs_vid') == 'true',
                'sw_ver': data.get('sfs_sw_ver') == 'true',
            }
        )

        return JsonResponse({'success': True, 'created': created})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})



def new_eng_telemetry(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)
    qr_code = cabinet.qr

    # =========================
    # 1️⃣ Новый способ через parse_device_status
    # =========================
    t1 = t2 = t3 = t4 = ''
    out1_val = ''

    if cabinet.iot_imei_locker:
        rssi, door_state, temperatures, out1_val = parse_device_status(cabinet.iot_imei_locker)
        if temperatures:
            if len(temperatures) > 0:
                t1 = str(temperatures[0])
            if len(temperatures) > 1:
                t2 = str(temperatures[1])

    # =========================
    # 2️⃣ Старый HTML парсер (fallback), если температур нет
    # =========================
    if not t1 and not t2:
        url = f'http://172.17.0.1:9000/detailed/{cabinet.iot_imei_locker}'
        try:
            resp = requests.get(
                url,
                auth=HTTPBasicAuth(settings.TELEMETRY_USER, settings.TELEMETRY_PASS),
                timeout=5,
            )
            resp.raise_for_status()
        except requests.RequestException as e:
            return render(request, 'ss_main/telemetry_partial.html', {
                'box_number': cabinet.shkaf_id,
                'error': f'Не удалось загрузить данные: {e}',
            })

        soup = BeautifulSoup(resp.content, 'html.parser')
        try:
            gen, tel, em = soup.find_all('div', class_='panel')
        except ValueError:
            return render(request, 'ss_main/telemetry_partial.html', {
                'box_number': cabinet.shkaf_id,
                'error': 'Ошибка структуры данных: ожидалось 3 блока .panel',
            })

        # --- Температуры как раньше ---
        temps = {}
        for tag in tel.find_all('p'):
            m = re.search(r'Температура\((\d)\).+?:\s*([\d.]+)', tag.get_text(strip=True))
            if m:
                temps[f't{m.group(1)}'] = m.group(2)

        t1 = temps.get('t1', '')
        t2 = temps.get('t2', '')
        t3 = temps.get('t3', '')
        t4 = temps.get('t4', '')

        # --- Остальные данные ---
        imei_text = gen.find('p').get_text(strip=True) if gen.find('p') else ''
        imei = imei_text.split(':', 1)[1].strip() if ':' in imei_text else 'n/a'

        iccid_a = gen.find('a', href=re.compile(r'/command/.*/iccid'))
        iccid = iccid_a['href'] if iccid_a else '-'

        def safe_split(p_tags, index):
            try:
                text = p_tags[index].get_text(strip=True)
                return text.split(':', 1)[1].strip() if ':' in text else 'n/a'
            except (IndexError, AttributeError):
                return 'n/a'

        p_tags = tel.find_all('p')
        last_update = safe_split(p_tags, 0)
        power_voltage = safe_split(p_tags, 2)
        reserv_voltage = power_voltage

        in_vals = {}
        for i in (1, 2, 3):
            tag = tel.find('p', text=re.compile(fr'Вход {i}:'))
            in_vals[f'in{i}'] = tag.get_text(strip=True).split(':', 1)[1].strip() if tag and ':' in tag.get_text() else ''

        coord_a = tel.find('a', href=re.compile(r'https?://'))
        try:
            coordinates = coord_a.get_text(strip=True).split(':', 1)[1].strip() if coord_a and ':' in coord_a.get_text() else ''
        except Exception:
            coordinates = ''

        gps_parts = [a.get_text(strip=True) for a in tel.find_all('a')
                     if 'GPS:' in a.get_text() or 'спутников' in a.get_text()]
        gps_info = ' '.join(gps_parts)

        h2_tag = em.find('h2')
        meter_reading = h2_tag.get_text(strip=True).split(':', 1)[1].strip() if h2_tag and ':' in h2_tag.get_text() else 'n/a'

        mp = em.find_all('p')
        meter_update = safe_split(mp, 0)
        meter_temp = safe_split(mp, 1)
        frequency = safe_split(mp, 2)

        table = em.find('table')
        phases = []
        if table:
            try:
                header_row = table.select('tr:first-of-type th')
                phases = [th.get_text(strip=True) for th in header_row[1:]]
            except Exception:
                phases = []

        def get_metric_values(metric_name):
            if not table or not phases:
                return ['0'] * len(phases)
            cell = table.find('td', text=re.compile(fr'^{re.escape(metric_name)}$'))
            if not cell:
                return ['0'] * len(phases)
            return [sib.get_text(strip=True) for sib in cell.find_next_siblings('td')]

        volt_vals = get_metric_values('Напряжение')
        curr_vals = get_metric_values('Ток')
        cos_vals = get_metric_values('Косинус φ')
        raw_power_vals = get_metric_values('Активная мощность')
        power_vals = get_metric_values('Активная мощность(расчёт)') if all(v in ('0', '') for v in raw_power_vals) else raw_power_vals

        def extract_num(s: str) -> str:
            m = re.search(r'-?\d+(?:\.\d+)?', s)
            return m.group(0) if m else '0'

        lines = []
        for idx, phase in enumerate(phases):
            v = extract_num(volt_vals[idx]) if idx < len(volt_vals) else '0'
            a = extract_num(curr_vals[idx]) if idx < len(curr_vals) else '0'
            cos = extract_num(cos_vals[idx]) if idx < len(cos_vals) else '0'
            p = extract_num(power_vals[idx]) if idx < len(power_vals) else '0'
            lines.append({
                'name': f'Line {phase}',
                'V': v,
                'A': a,
                'cos': cos,
                'P': p,
            })

        context = {
            'box_number':    cabinet.shkaf_id,
            'qr_code':       qr_code,
            'imei':          imei,
            'iccid':         cabinet.iot_imei_locker,
            'last_update':    last_update,
            'reserv_voltage': reserv_voltage,
            'power_voltage':  power_voltage,
            **in_vals,
            'out1':           out1_val,
            'out2':           '',
            't1':             t1,
            't2':             t2,
            't3':             t3,
            't4':             t4,
            'coordinates':    coordinates,
            'gps_info':       gps_info,
            'power_count':   meter_reading,
            'meter_update':  meter_update,
            'meter_temp':    meter_temp,
            'frequency':     frequency,
            'lines':         lines,
            'sticker':       getattr(cabinet, 'sticker', ''),
        }
        return render(request, 'ss_main/telemetry_partial.html', context)

    # =========================
    # 3️⃣ Если данные успешно получены из parse_device_status
    # =========================
    context = {
        'box_number': cabinet.shkaf_id,
        'qr_code': qr_code,
        'imei': cabinet.iot_imei_locker or 'n/a',
        'iccid': cabinet.iot_imei_locker or 'n/a',
        't1': t1,
        't2': t2,
        't3': t3,
        't4': t4,
        'out1': out1_val,
        'out2': '',
        'last_update': 'n/a',
        'reserv_voltage': 'n/a',
        'power_voltage': 'n/a',
        'in1': '', 'in2': '', 'in3': '',
        'coordinates': '',
        'gps_info': '',
        'power_count': '',
        'meter_update': '',
        'meter_temp': '',
        'frequency': '',
        'lines': [],
        'sticker': getattr(cabinet, 'sticker', ''),
    }
    return render(request, 'ss_main/telemetry_partial.html', context)



def update_sticker(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)
    new_sticker = request.POST.get('sticker', '').strip()
    cabinet.sticker = new_sticker
    cabinet.save(update_fields=['sticker'])
    return JsonResponse({'success': True, 'sticker': cabinet.sticker})


@user_passes_test(is_engineer)
def cabinet_settings(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)
    settings, _ = Cabinet_settings_for_auto_marking.objects.get_or_create(cabinet_id=cabinet)
    settings_for_settings = Settings_for_settings.objects.filter(settings_for=settings).first()



    if request.method == 'POST':
        form = CabinetSettingsForm(request.POST, instance=settings)
        settings_for_form = SettingsForSettingsForm(request.POST, instance=settings_for_settings)

        if form.is_valid() and settings_for_form.is_valid():
            form.save()
            settings_for_form.save()

            Cell.objects.filter(cabinet_id=cabinet).update(is_error=False, message=None)

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': {**form.errors, **settings_for_form.errors}})

    form = CabinetSettingsForm(instance=settings)
    settings_for_form = SettingsForSettingsForm(instance=settings_for_settings)

    return render(request, 'ss_main/cabinet_settings_partial.html', {
        'form': form,
        'settings_for_form': settings_for_form,
        'cabinet': cabinet
    })



@user_passes_test(is_engineer)
def cabinet_settings2(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id)
    settings, _ = Cabinet_settings_for_auto_marking.objects.get_or_create(cabinet_id=cabinet)
    settings_for_settings = Settings_for_settings.objects.filter(settings_for=settings).first()

    if request.method == 'POST':
        form = CabinetSettingsForm(request.POST, instance=settings)
        settings_for_form = SettingsForSettingsForm(request.POST, instance=settings_for_settings)

        if form.is_valid() and settings_for_form.is_valid():
            form.save()
            settings_for_form.save()

            # Обновляем статусы ячеек
            Cell.objects.filter(cabinet_id=cabinet).update(is_error=False, message=None)

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': {**form.errors, **settings_for_form.errors}})

    form = CabinetSettingsForm(instance=settings)
    settings_for_form = SettingsForSettingsForm(instance=settings_for_settings)

    return render(request, 'ss_main/cabinet_settings_partial2.html', {
        'form': form,
        'settings_for_form': settings_for_form,
        'cabinet': cabinet
    })


@user_passes_test(is_engineer)
def main(request):
    cities = Cabinet.objects.values_list('city', flat=True).distinct()
    return render(request, 'ss_main/index.html', {'cities': cities})


@login_required
@user_passes_test(is_logistician)
def create_courier(request):
    if request.method == 'POST':
        form = CourierCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(10))
            user.set_password(password)
            user.username = f'{user.first_name}_{user.last_name}'
            user.save()

            current_user_citys = request.user.citys.all()
            for city in current_user_citys:
                user.citys.add(city)

            send_mail(
                'Scout Registration Info',
                f'Username: {user.username}\nPassword: {password}\nEmail: {user.email}\nFull Name: {user.first_name} {user.last_name}',
                'bhromenko@mail.ru',
                [request.user.email, user.email],
                fail_silently=False,
            )
            messages.success(request, f"Скаут '{user.username}' успешно зарегистрирован (проверьте свою почту).")
            return redirect('create_courier')
        else:
            messages.error(request, "Пожалуйста, исправьте ошибки в форме.")
    else:
        form = CourierCreationForm()

    return render(request, 'ss_main/create_courier.html', {'form': form})



# Создание нового логиста
@login_required
@user_passes_test(is_regional_manager)
def create_logic(request):
    if request.method == 'POST':
        form = LogicCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(10))
            user.set_password(password)
            user.username = f'{user.first_name}_{user.last_name}'
            user.save()

            send_mail(
                'Logistician Registration Info',
                f'Username: {user.username}\nPassword: {password}\nEmail: {user.email}\nFull Name: {user.first_name} {user.last_name}',
                'bhromenko@mail.ru',
                [request.user.email, user.email],
                fail_silently=False,
            )
            messages.success(request, f"Логист '{user.username}' успешно зарегистрирован (проверьте свою почту).")
    else:
        form = LogicCreationForm()

    return render(request, 'ss_main/create_logic.html', {'form': form})


# Назначение зоны курьеру
@login_required
@user_passes_test(is_logistician)
def assign_zone_to_courier(request):
    user_city = request.user.citys.first()
    if request.method == 'POST':
        courier_id = request.POST.get('courier_id')
        zone_id = request.POST.get('zone_id')
        try:
            courier = get_object_or_404(CustomUser, pk=courier_id)
            zone = get_object_or_404(Zone, pk=zone_id)
            if zone.city == user_city:
                courier.zones.clear()
                courier.zones.add(zone)
                messages.success(request, f"Zone '{zone.zone_name}' successfully assigned to courier '{courier.username}'.")
            else:
                messages.error(request, "The selected zone is not in your city.")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('assign_zone_to_courier')

    couriers = CustomUser.objects.filter(role='courier', citys=user_city)
    zones = Zone.objects.filter(city=user_city)
    return render(request, 'ss_main/assign_zone.html', {'couriers': couriers, 'zones': zones})


# Удаление курьера
@login_required
@user_passes_test(is_logistician)
def delete_courier(request, courier_id):
    courier = get_object_or_404(CustomUser, pk=courier_id)
    courier.delete()
    return JsonResponse({'status': 'success'})


@login_required
@user_passes_test(is_regional_manager)
def assign_zone_to_logic(request):
    if request.method == 'POST':
        logistician_id = request.POST.get('logistician_id')
        zone_id = request.POST.get('zone_id')

        try:
            logistician = get_object_or_404(CustomUser, pk=logistician_id)
            zone = get_object_or_404(Zone, pk=zone_id)

            logistician.zones.clear()
            logistician.zones.add(zone)
            messages.success(request, f"Zone '{zone.zone_name}' successfully assigned to logistician '{logistician.username}'.")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('assign_zone_to_logic')

    logisticians = CustomUser.objects.filter(role='logistician')
    zones = Zone.objects.all()
    return render(request, 'ss_main/assign_logic.html', {'logisticians': logisticians, 'zones': zones})


# Список зон (основная страница логиста)
@login_required
@user_passes_test(is_logistician)
def zone_list(request):
    user = request.user
    user_city = user.citys.first()
    if user_city:
        zones = Zone.objects.filter(city=user_city)
    else:
        zones = Zone.objects.none()

    zone_data = []

    for zone in zones:
        cabinets = Cabinet.objects.filter(zone=zone)
        capacity_sum = 0
        buffer_sum = 0
        locked_doors_count = 0

        for cabinet in cabinets:
            try:
                capacity_sum += int(cabinet.capacity or 0)
                buffer_sum += int(cabinet.buffer or 0)
            except ValueError:
                pass  # игнорируем ошибки преобразования

            if cabinet.door_state:  # True — дверь закрыта
                locked_doors_count += 1

        zone_data.append({
            'zone': zone,
            'cabinet_count': cabinets.count(),
            'capacity_sum': capacity_sum,
            'buffer_sum': buffer_sum,
            'locked_doors_count': locked_doors_count
        })

    cabinets_with_coords = Cabinet.objects.filter(latitude__isnull=False, longitude__isnull=False)
    cabinets_json = json.dumps(
        list(cabinets_with_coords.values('shkaf_id', 'street', 'latitude', 'longitude')),
        cls=DjangoJSONEncoder
    )

    return render(request, 'ss_main/zone_list.html', {
        'zone_data': zone_data,
        'user_city': user_city,
        'cabinets_with_coords': cabinets_json
    })


@login_required
@user_passes_test(is_logistician)
def zone_detail(request, zone_id):
    zone = get_object_or_404(Zone, id=zone_id)
    couriers = CustomUser.objects.filter(zones=zone, role='courier')
    return render(request, 'ss_main/zone_detail.html', {'zone': zone, 'couriers': couriers})



# переехать до конца на это
def parse_device_status(iot_imei_locker, devices_cache=None):
    """
    Возвращает:
        rssi_signal: str
        door_state: bool | None
        temperatures: list[float] | None
        out1_val: str ("LOCKED"/"OPEN"/"")
    """
    try:
        if not iot_imei_locker:
            return 'n/a', None, None, ''

        # если кэш не передан — грузим сами (fallback)
        if devices_cache is None:
            response = requests.get("http://172.17.0.1:9001/api/cache", timeout=3)
            response.raise_for_status()
            devices_cache = response.json()

        # --- Ищем устройство по IMEI ---
        target = next((dev for dev in devices_cache if str(dev.get("IMEI")) == str(iot_imei_locker)), None)
        if not target:
            return 'нет данных', None, None, ''

        # --- RSSI ---
        gsm = target.get("GSMStatus")
        if gsm is None:
            rssi_signal = 'n/a'
        elif gsm == 31:
            rssi_signal = '-51 дБм ++'
        elif gsm == 0:
            rssi_signal = '-113 дБм --'
        elif 0 < gsm <= 30:
            rssi_signal = f'{-113 + gsm * 2} дБм'
        elif gsm == 99:
            rssi_signal = 'нет сигнала'
        else:
            rssi_signal = 'ошибка'

        # --- Door state + OUT1 ---
        door_state = None
        out1_val = ''

        if "is_door_locked" in target and isinstance(target["is_door_locked"], list):
            locked = target["is_door_locked"][0]
            # Инвертируем значение
            door_state = not locked
            out1_val = "LOCKED" if door_state else "OPEN"
        else:
            digital_in1 = target.get("DigitalIn1")
            if digital_in1 is not None:
                exhaust_bit = (digital_in1 // 2) % 2
                # Инвертируем значение
                door_state = not (exhaust_bit == 0)
                out1_val = "LOCKED" if door_state else "OPEN"

        # --- Temperature ---
        temp1 = target.get("Temperature1_1wire")
        temp2 = target.get("Temperature2_1wire")

        temperatures = []
        for t in (temp1, temp2):
            if t is not None and t != -128:
                temperatures.append(t)

        if not temperatures:
            temperatures = None

        return rssi_signal, door_state, temperatures, out1_val

    except Exception:
        return 'ошибка', None, None, ''




@login_required
@user_passes_test(is_logistician)
def cabinet_list(request, zone_id):
    zone = get_object_or_404(Zone, pk=zone_id)
    cabinets = Cabinet.objects.filter(zone=zone)
    cabinets_count = cabinets.count()
    cabinets_with_coords = list(cabinets.values('shkaf_id', 'street', 'latitude', 'longitude'))

    total_capacity = 0
    cabinets_status_counts = []

    for cabinet in cabinets:
        # Суммируем capacity
        try:
            capacity_value = int(cabinet.capacity) if cabinet.capacity is not None else 0
            total_capacity += capacity_value
        except (ValueError, TypeError):
            pass

        # Получаем связанные ячейки
        cells = Cell.objects.filter(cabinet_id=cabinet)

        status_counts = {
            'ready': cells.filter(status='ready').count(),
            'charging': cells.filter(status='charging').count(),
            'empty': cells.filter(status='empty').count(),
            'Inactive': cells.filter(status='Inactive').count(),
            'ban': cells.filter(status='BAN').count(),
        }

        rssi_signal, door_state, _, _ = parse_device_status(cabinet.iot_imei_locker)

        updated_fields = []
        if cabinet.rssi != rssi_signal:
            cabinet.rssi = rssi_signal
            updated_fields.append('rssi')

        if door_state is not None and cabinet.door_state != door_state:
            cabinet.door_state = door_state
            updated_fields.append('door_state')

        if updated_fields:
            cabinet.save(update_fields=updated_fields)

        cabinets_status_counts.append((cabinet, status_counts, rssi_signal))

    return render(request, 'ss_main/cabinet_list.html', {
        'zone': zone,
        'cabinets_count': cabinets_count,
        'total_cells_count': total_capacity,
        'cabinets_status_counts': cabinets_status_counts,
        'cabinets_with_coords': cabinets_with_coords,
    })


@login_required
@user_passes_test(is_regional_manager)
def region_cabinet_list(request, zone_id):
    zone = get_object_or_404(Zone, pk=zone_id)
    cabinets = Cabinet.objects.filter(zone=zone)
    cabinets_count = cabinets.count()

    total_capacity = 0
    cabinets_status_counts = []

    for cabinet in cabinets:
        # 📌 Суммируем capacity
        try:
            capacity_value = int(cabinet.capacity) if cabinet.capacity else 0
            total_capacity += capacity_value
        except (ValueError, TypeError):
            pass

        # 📌 Получаем связанные ячейки
        cells = Cell.objects.filter(cabinet_id=cabinet)
        status_counts = {
            'ready': cells.filter(status='ready').count(),
            'charging': cells.filter(status='charging').count(),
            'empty': cells.filter(status='empty').count(),
            'Inactive': cells.filter(status='Inactive').count(),
            'ban': cells.filter(status='BAN').count(),
        }

        # 📌 Парсим RSSI, состояние двери и температуру из кэша (parse_device_status)
        rssi_signal, door_state, temperatures, _ = parse_device_status(cabinet.iot_imei_locker)
        updated_fields = []

        if cabinet.rssi != rssi_signal:
            cabinet.rssi = rssi_signal
            updated_fields.append('rssi')

        if door_state is not None and cabinet.door_state != door_state:
            cabinet.door_state = door_state
            updated_fields.append('door_state')

        # 📌 Температура из кэша
        temperature = '-'
        if temperatures and len(temperatures) > 0:
            temperature = str(temperatures[0])
            if cabinet.temperature1 != temperature:
                cabinet.temperature1 = temperature
                updated_fields.append('temperature1')

        # 📌 Power Count и Grid Voltage — по старому HTML способу
        power_count, grid_voltage = '-', '-'
        try:
            if cabinet.iot_imei_locker:
                url = f'http://172.17.0.1:9000/detailed/{cabinet.iot_imei_locker}'
                resp = requests.get(
                    url,
                    auth=HTTPBasicAuth(settings.TELEMETRY_USER, settings.TELEMETRY_PASS),
                    timeout=3
                )
                resp.raise_for_status()
                soup = BeautifulSoup(resp.content, 'html.parser')

                gen, tel, em = soup.find_all('div', class_='panel')

                # Power Count
                try:
                    h2_tag = em.find('h2')
                    if h2_tag and ':' in h2_tag.text:
                        power_count = h2_tag.text.split(':', 1)[1].strip()
                except Exception:
                    power_count = '-'

                # Grid Voltage
                def extract_num(s: str) -> str:
                    m = re.search(r'-?\d+(?:\.\d+)?', s)
                    return m.group(0) if m else '0'

                table = em.find('table')
                volt_vals = []
                if table:
                    cell = table.find('td', text=re.compile(r'^Напряжение$'))
                    if cell:
                        volt_vals = [extract_num(td.get_text(strip=True)) for td in cell.find_next_siblings('td')]
                grid_voltage = '/ '.join(volt_vals) if volt_vals else '-'

        except Exception as e:
            print(f"[ERROR] Cabinet {cabinet.shkaf_id}: {e}")

        # 📌 Сохраняем обновлённые поля
        if power_count != '-' and cabinet.power_count != power_count:
            cabinet.power_count = power_count
            updated_fields.append('power_count')

        if grid_voltage != '-' and cabinet.grid_voltage != grid_voltage:
            cabinet.grid_voltage = grid_voltage
            updated_fields.append('grid_voltage')

        if updated_fields:
            cabinet.save(update_fields=updated_fields)

        cabinets_status_counts.append(
            (cabinet, status_counts, rssi_signal, power_count, grid_voltage, temperature)
        )

    cabinets_with_coords = list(cabinets.values('shkaf_id', 'street', 'latitude', 'longitude'))

    return render(request, 'ss_main/region_cabinets.html', {
        'zone': zone,
        'cabinets_count': cabinets_count,
        'total_cells_count': total_capacity,
        'cabinets_status_counts': cabinets_status_counts,
        'cabinets_with_coords': cabinets_with_coords,
    })


def cabinet_details_api(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id, zone__users=request.user)
    cells = Cell.objects.filter(cabinet_id=cabinet)

    cabinet_data = {
        'shkaf_id': cabinet.shkaf_id,
        'city': {'city_name': cabinet.city.city_name},
        'zone': {'zone_name': cabinet.zone.zone_name},
        'street': cabinet.street,
        'location': cabinet.location,
        'extra_inf': cabinet.extra_inf,
        'cells': [
            {'endpointid': cell.endpointid, 'cap_percent': cell.cap_percent, 'vid': cell.vid}
            for cell in cells
        ]
    }

    return JsonResponse(cabinet_data)


@login_required
@user_passes_test(is_regional_manager)
def main_region(request):
    cities = City.objects.all()
    city_data = []
    for city in cities:
        zones_count = Zone.objects.filter(city=city).count()
        cabinets = Cabinet.objects.filter(city=city)

        cabinets_count = cabinets.count()

        total_capacity = 0
        total_buffer = 0
        for cab in cabinets:
            try:
                total_capacity += int(cab.capacity or 0)
            except (TypeError, ValueError):
                continue
            try:
                total_buffer += int(cab.buffer or 0)
            except (TypeError, ValueError):
                continue

        city_data.append({
            'city': city,
            'zones_count': zones_count,
            'cabinets_count': cabinets_count,
            'charging_capacity': total_capacity,
            'buffer_capacity': total_buffer,
        })

    return render(request, 'ss_main/main_region.html', {
        'city_data': city_data,
        'cities': cities
    })


@login_required
@user_passes_test(is_regional_manager)
def region_zones(request, city_id):
    city = get_object_or_404(City, id=city_id)
    zones = Zone.objects.filter(city=city)
    zone_data = []
    total_cells = 0
    total_cabinets = 0

    for zone in zones:
        status_counts = zone.cell_status_counts()
        total_cells += sum(status_counts.values())

        cabinets = Cabinet.objects.filter(zone=zone)
        cabinets_count = cabinets.count()
        total_cabinets += cabinets_count

        couriers_count = CustomUser.objects.filter(zones=zone, role='courier').count()
        logisticians_count = CustomUser.objects.filter(zones=zone, role='logistician').count()

        # Суммируем capacity
        total_capacity = 0
        for cab in cabinets:
            try:
                total_capacity += int(cab.capacity)
            except (TypeError, ValueError):
                continue

        # Суммируем buffer
        total_buffer = 0
        for cab in cabinets:
            try:
                total_buffer += int(cab.buffer)
            except (TypeError, ValueError):
                continue

        # Суммируем power_count
        total_power_count = 0
        for cab in cabinets:
            if cab.power_count:
                match = re.search(r"[-+]?\d*\.\d+|\d+", cab.power_count)
                if match:
                    try:
                        total_power_count += float(match.group(0))
                    except ValueError:
                        continue

        zone_data.append({
            'zone': zone,
            'status_counts': status_counts,
            'cabinets_count': cabinets_count,
            'couriers_count': couriers_count,
            'logisticians_count': logisticians_count,
            'charging_capacity': round(total_capacity, 2),
            'buffer_capacity': round(total_buffer, 2),
            'sum_power_count': round(total_power_count, 2),
        })

    cabinets_with_coords = Cabinet.objects.filter(latitude__isnull=False, longitude__isnull=False)
    cabinets_json = json.dumps(
        list(cabinets_with_coords.values('shkaf_id', 'street', 'latitude', 'longitude')),
        cls=DjangoJSONEncoder
    )

    return render(request, 'ss_main/region_zones.html', {
        'zone_data': zone_data,
        'city': city,
        'total_zones': len(zone_data),
        'total_cells': total_cells,
        'cabinets_with_coords': cabinets_json,
        'total_cabinets': total_cabinets,
    })



# Логисты зоны(региональный менеджер)
@login_required
@user_passes_test(is_regional_manager)
def region_logic(request):

    logisticians = CustomUser.objects.filter(role='logistician')

    return render(request, 'ss_main/region_logic.html', {
        'logisticians': logisticians,
    })



# Основная страница курьера
@login_required
@user_passes_test(is_courier)
def user_cabinets(request):
    user = request.user
    zones = user.zones.all()
    cabinets = Cabinet.objects.filter(zone__in=zones)

    cabinet_statuses = []
    for cabinet in cabinets:
        cells = cabinet.cell_set.all()
        ready_count = cells.filter(status='ready').count()
        charging_count = cells.filter(status='charging').count()
        empty_count = cells.filter(status='empty').count()
        ban_count = cells.filter(status='BAN').count()
        cabinet_statuses.append({
            'cabinet': cabinet,
            'ready_count': ready_count,
            'charging_count': charging_count,
            'empty_count': empty_count,
            'ban_count': ban_count

        })

    context = {
        'cabinet_statuses': cabinet_statuses,
        'user': user,
        'zones': zones,
    }
    return render(request, 'ss_main/user_cabinets.html', context)



@login_required
def cabinet_details(request, shkaf_id):
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id, zone__users=request.user)
    cells = Cell.objects.filter(cabinet_id=cabinet)
    cabinet_setting = Cabinet_settings_for_auto_marking.objects.filter(cabinet_id=cabinet.shkaf_id).first()
    critical_temp = cabinet_setting.critical_temp if cabinet_setting else None
    status_counts = cells.values('status').annotate(count=Count('status')).order_by('status')
    status_slots = {}

    for cell in cells:
        status = cell.status
        temp_cur1 = cell.temp_cur1
        endpointid = cell.endpointid
        sw_name = cell.sw_name
        cap_percent = cell.cap_percent or "N/A"
        if status not in status_slots:
            status_slots[status] = []
        status_slots[status].append({'endpointid': endpointid, 'charge': cap_percent, 'sw_name': sw_name, 'temp_cur1': temp_cur1,})

    average_charge = cells.annotate(cap_percent_as_float=Cast('cap_percent', FloatField())).aggregate(
        average_charge=Avg('cap_percent_as_float'))['average_charge']

    error_slots = cells.filter(is_error=True)

    # --- Временные интервалы ---
    almaty = pytz.timezone('Asia/Almaty')
    now = datetime.datetime.now(almaty)

    today_9 = now.replace(hour=9, minute=0, second=0, microsecond=0)
    today_21 = now.replace(hour=21, minute=0, second=0, microsecond=0)

    # По какому интервалу показывать first_half / second_half
    if now < today_9:
        # Сейчас ночь: показать итог с 21:00 вчера до 09:00 сегодня
        target_date = (now - datetime.timedelta(days=1)).date()
        period = "night"
    elif now < today_21:
        # Сейчас день: показать итог с 09:00 до 21:00 сегодня
        target_date = now.date()
        period = "day"
    else:
        # Сейчас ночь: показать итог с 21:00 до 09:00 завтра
        target_date = now.date()
        period = "night"

    # Получаем нужную запись истории
    history_entry = Cabinet_history.objects.filter(history_for=cabinet, date=target_date).first()

    # Определяем, какие данные использовать
    if period == "day":
        first_half_count = history_entry.first_half if history_entry else 0
        second_half_count = 0  # предыдущая ночь нас не интересует
    else:
        first_half_count = history_entry.second_half if history_entry else 0
        second_half_count = 0  # предыдущий день не нужен

    # Получаем "last full day" — интервал 09:00 до 09:00
    # Берем вчерашнюю дату
    full_day_date = (now - datetime.timedelta(days=1)).date()
    full_day_entry = Cabinet_history.objects.filter(history_for=cabinet, date=full_day_date).first()
    full_day_count = 0
    if full_day_entry:
        full_day_count = (full_day_entry.first_half or 0) + (full_day_entry.second_half or 0)

    context = {
        'cabinet': cabinet,
        'status_counts': json.dumps(list(status_counts)),
        'status_slots': json.dumps(status_slots),
        'average_charge': average_charge,
        'error_slots': error_slots,
        'first_half_count': first_half_count,
        'second_half_count': second_half_count,
        'full_day_count': full_day_count,
        'latitude': cabinet.latitude,
        'longitude': cabinet.longitude,
        'critical_temp': critical_temp,
    }
    return render(request, 'ss_main/scout_v2.html', context)


@login_required
def open_cabinet(request, shkaf_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Только POST запрос"}, status=405)

    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id, zone__users=request.user)

    Eventlogger.objects.create(
        user=request.user.username,
        shkaf_id=str(cabinet.shkaf_id),
        login=request.user.username,
        action=f"Open attempt for cabinet {cabinet.shkaf_id}",
        time=timezone.now(),
        door_state_before=getattr(cabinet, "door_state", None)
    )

    # Проверяем IMEI
    imei = cabinet.iot_imei_locker
    if not imei:
        return JsonResponse({"success": False, "message": "IMEI не задан"}, status=400)

    # Отправляем команду открытия
    url = f"http://172.17.0.1:9001/api/dev/{imei}/out?lineno=1&state=1"
    try:
        resp = requests.get(url, timeout=5)
        if resp.status_code == 200:
            return JsonResponse({"success": True, "message": "Команда на открытие отправлена ✅"})
        else:
            return JsonResponse({
                "success": False,
                "message": f"Ошибка от API: {resp.status_code}",
                "details": resp.text
            }, status=resp.status_code)
    except requests.RequestException as e:
        return JsonResponse({"success": False, "message": f"Ошибка соединения: {e}"}, status=500)




def export_battery_history(request, shkaf_id):
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    if not start_date_str or not end_date_str:
        return HttpResponse("Обе даты обязательны", status=400)

    try:
        # Преобразование строки в datetime с учётом часового пояса
        start_naive = datetime.datetime.strptime(start_date_str, "%Y-%m-%d")
        end_naive = datetime.datetime.strptime(end_date_str, "%Y-%m-%d")

        # Сделать aware
        start_date = timezone.make_aware(start_naive)
        # Чтобы включить весь день до конца, добавим один день и уберем микросекунду
        end_date = timezone.make_aware(end_naive + datetime.timedelta(days=1)) - datetime.timedelta(microseconds=1)
    except Exception as e:
        return HttpResponse(f"Неверный формат даты: {e}", status=400)

    try:
        history = Cabinet_history.objects.filter(
            history_for__shkaf_id=shkaf_id,
            date__range=(start_date, end_date)
        ).order_by("date")
    except Exception as e:
        return HttpResponse(f"Ошибка получения данных: {e}", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "History"

    ws.append(["Дата", "FIRST_HALF", "SECOND_HALF", "first_data", "second_data"])

    for entry in history:
        ws.append([
            timezone.localtime(entry.date).strftime("%Y-%m-%d %H:%M:%S") if entry.date else "",
            entry.first_half,
            entry.second_half,
            entry.first_data or "",
            entry.second_data or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"cabinet_{shkaf_id}_history.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def update_cabinet_data(request, shkaf_id):
    logger.debug(f"Headers: {request.headers}")
    logger.debug(f"Method: {request.method}")
    cabinet = get_object_or_404(Cabinet, shkaf_id=shkaf_id, zone__users=request.user)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        cells = Cell.objects.filter(cabinet_id=cabinet)

        status_counts = cells.values('status').annotate(count=Count('status'))
        average_charge = cells.annotate(cap_percent_as_float=Cast('cap_percent', FloatField())).aggregate(
            average_charge=Avg('cap_percent_as_float'))['average_charge']

        status_slots = {}
        for cell in cells:
            status = cell.status
            if status not in status_slots:
                status_slots[status] = []
            status_slots[status].append({'endpointid': cell.endpointid, 'charge': cell.cap_percent, 'sw_name': cell.sw_name, 'temp_cur1': cell.temp_cur1,})
        error_slots = cells.filter(is_error=True)
        error_slots_list = [{'endpointid': slot.endpointid, 'message': slot.message} for slot in error_slots]
        almaty_timezone = pytz.timezone('Asia/Almaty')
        current_time = datetime.datetime.now(almaty_timezone).replace(tzinfo=None)
        history_date = current_time.date()
        history_entry = Cabinet_history.objects.filter(history_for=cabinet, date=history_date).first()
        first_half_count = history_entry.first_half if history_entry else 0
        second_half_count = history_entry.second_half if history_entry else 0
        return JsonResponse({
            'status_counts': list(status_counts),
            'status_slots': status_slots,
            'average_charge': average_charge,
            'error_slots': error_slots_list,
            'first_half_count': first_half_count,
            'second_half_count': second_half_count,
            'latitude': cabinet.latitude,
            'longitude': cabinet.longitude,
        })
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)



@login_required
def user_cabinets_api(request):
    user = request.user
    zones = user.zones.all()
    cabinets = Cabinet.objects.filter(zone__in=zones)

    cabinet_statuses = []
    for cabinet in cabinets:
        cells = cabinet.cell_set.all()
        ready_count = cells.filter(status='ready').count()
        charging_count = cells.filter(status='charging').count()
        empty_count = cells.filter(status='empty').count()
        ban_count = cells.filter(status='BAN').count()
        cabinet_statuses.append({
            'shkaf_id': cabinet.shkaf_id,
            "readable_name": cabinet.readable_name,
            'ready_count': ready_count,
            'charging_count': charging_count,
            'empty_count': empty_count,
            'ban_count': ban_count,
        })

    return JsonResponse(cabinet_statuses, safe=False)



def station_ids(request):
    query = request.GET.get('query', '')
    data = Report.objects.using('new_db').filter(stationid__icontains=query)
    suggestions = list(data.values_list('stationid', flat=True).distinct())
    return JsonResponse(suggestions, safe=False)


def cities(request):
    query = request.GET.get('query', '')
    data = Report.objects.using('new_db').filter(city__icontains=query)
    suggestions = list(data.values_list('city', flat=True).distinct())
    return JsonResponse(suggestions, safe=False)



def zones(request):
    query = request.GET.get('query', '')
    data = Report.objects.using('new_db').filter(zone__icontains=query)
    suggestions = list(data.values_list('zone', flat=True).distinct())
    return JsonResponse(suggestions, safe=False)


@staff_required
def report(request):
    if request.method == 'POST':
        form = ReportFilterForm(request.POST)
        if form.is_valid():
            station_id = form.cleaned_data.get('station_id')
            city = form.cleaned_data.get('city')
            zone = form.cleaned_data.get('zone')
            time_from = form.cleaned_data.get('time_from')
            time_to = form.cleaned_data.get('time_to')

            if not all([station_id, city, zone, time_from, time_to]):
                return HttpResponse("Пожалуйста, выберитpе все фильтры.")

            reports = Report.objects.using('new_db').all()

            if station_id:
                reports = reports.filter(stationid__startswith=station_id)
            if city:
                reports = reports.filter(city=city)
            if zone:
                reports = reports.filter(zone=zone)
            if time_from and time_to:
                time_to += timedelta(days=1)
                reports = reports.filter(time__range=[time_from, time_to])

                reports = reports.filter(time__range=[time_from, time_to])

            if reports.exists():
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="reports.xlsx"'

                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                headers = ['city', 'zone', 'stationid', 'reason', 'balance_status', 'capacity', 'cap_coulo',
                           'cap_percent', 'cap_vol', 'charge_cap_h', 'charge_cap_l', 'charge_times', 'core_volt',
                           'current_cur', 'cycle_times', 'design_voltage', 'fun_boolean', 'healthy', 'ochg_state',
                           'odis_state', 'over_discharge_times', 'pcb_ver', 'remaining_cap', 'remaining_cap_percent',
                           'sn', 'sw_ver', 'temp_cur1', 'temp_cur2', 'total_capacity', 'vid', 'voltage_cur',
                           'session_start', "start_percent", 'time']
                worksheet.append(headers)


                for report in reports:
                    row = [report.city, report.zone, report.stationid, report.reason, report.balance_status,
                           report.capacity, report.cap_coulo, report.cap_percent, report.cap_vol, report.charge_cap_h,
                           report.charge_cap_l, report.charge_times, report.core_volt, report.current_cur,
                           report.cycle_times, report.design_voltage, report.fun_boolean, report.healthy,
                           report.ochg_state, report.odis_state, report.over_discharge_times, report.pcb_ver,
                           report.remaining_cap, report.remaining_cap_percent, report.sn, report.sw_ver,
                           report.temp_cur1, report.temp_cur2, report.total_capacity, report.vid, report.voltage_cur,
                           report.session_start.replace(tzinfo=None) if report.session_start else None,
                           report.start_percent,
                           report.time.replace(tzinfo=None) if report.time else None]
                    worksheet.append(row)


                workbook.save(response)
                return response
            else:
                return HttpResponse("Для выбранной зарядной станции, города или зоны отчеты отсутствуют.")
    else:
        form = ReportFilterForm()
        prefixes = Report.objects.using('new_db').values_list('stationid', flat=True).distinct()
        prefixes = set(prefix.split('-')[0] for prefix in prefixes)
        cities = Report.objects.using('new_db').values_list('city', flat=True).distinct()
        zones = Report.objects.using('new_db').values_list('zone', flat=True).distinct()

        return render(request, 'ss_main/report.html', {
            'form': form,
            'station_ids': list(prefixes),
            'cities': list(cities),
            'zones': list(zones)
        })


def reset_selection(request):
    if request.method == 'GET':
        if 'station_id' in request.session:
            del request.session['station_id']
        if 'select_all' in request.session:
            del request.session['select_all']
    return redirect('report')


def report_page(request):
    cabinets = Eventlogger.objects.values_list("shkaf_id", flat=True).distinct()
    return render(request, "ss_main/logreports.html", {"cabinets": cabinets})


def download_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    shkaf_id = request.GET.get("shkaf_id")
    report_type = request.GET.get("report_type")

    start_date = parse_date(start_date) if start_date else None
    end_date = parse_date(end_date) if end_date else None

    qs = Eventlogger.objects.all()

    if start_date:
        qs = qs.filter(time__date__gte=start_date)
    if end_date:
        qs = qs.filter(time__date__lte=end_date)
    if shkaf_id:
        qs = qs.filter(shkaf_id=shkaf_id)

    if report_type == "events":
        qs = qs.order_by("time")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["Время", "Шкаф", "Readable name", "Пользователь", "Login", "Действие", "Door state before"]
    ws.append(headers)

    for row in qs:
        ws.append([
            row.time.strftime("%Y-%m-%d %H:%M:%S") if row.time else "",
            row.shkaf_id,
            row.readable_name,
            row.user,
            row.login,
            row.action,
            "Открыта" if row.door_state_before else "Закрыта"
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename=report.xlsx'
    wb.save(response)
    return response


class CustomLoginView(LoginView):
    form_class = CustomAuthenticationForm
    template_name = 'ss_main/login.html'

    def get_success_url(self):
        user = self.request.user
        if user.role == 'courier':
            return '/my-cabinets/'
        elif user.role == 'logistician':
            return '/zones/'
        elif user.role == 'regional_manager':
            return '/region/'
        elif user.role == 'engineer':
            return '/'
        else:
            return '/'


class CustomLogoutView(LogoutView):
    next_page = '/'
    http_method_names = ['post']
